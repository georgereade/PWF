import requests
from requests.auth import HTTPBasicAuth
import os
from dotenv import load_dotenv
import pandas as pd
import time

from datetime import datetime, timedelta

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

load_dotenv()

# Define your API credentials
API_KEY = os.getenv('PWF_API_KEY')  # Replace with your actual ProWorkflow API key
BASE_URL = 'https://api.proworkflow.net'  # Base URL for the ProWorkflow API
USERNAME = os.getenv('PWF_USERNAME')  # Replace with your ProWorkflow username
PASSWORD = os.getenv('PWF_PASSWORD')  # Replace with your ProWorkflow password

# Headers for authentication
headers = {
    'Content-Type': 'application/json',
    'apikey': API_KEY,
}

pd.set_option('display.max_colwidth', None)

trackedfrom = '2024-01-01'  # Specify start date for the time range
trackedto = '2024-12-31'   # Specify end date for the time range

# Function to get all contacts of type 'staff'
def get_staff_contacts():
    url = f'{BASE_URL}/contacts'
    response = requests.get(url, headers=headers, auth=HTTPBasicAuth(USERNAME, PASSWORD))
    if response.status_code == 200:
        contacts = response.json()['contacts']
        print("Collected staff names...")
        return [contact for contact in contacts if contact['type'] == 'staff']
    else:
        raise Exception(f"Error fetching contacts: {response.status_code}, {response.text}")

# Second request: Function to get task-specific time details by contact
def get_contact_task_details(contact_id, trackedfrom, trackedto):
    url = f'{BASE_URL}/contacts/{contact_id}/time?trackedfrom={trackedfrom}&trackedto={trackedto}&fields=dates,project,task,notes,contact,category'
    response = requests.get(url, headers=headers, auth=HTTPBasicAuth(USERNAME, PASSWORD))
    if response.status_code == 200:
        task_times = response.json()['timerecords']
        filtered_records = [record for record in task_times if record['categoryname'] in ["On Hold", "Current Timed Projects"]]
        return filtered_records
    else:
        raise Exception(f"Error fetching task details for contact {contact_id}: {response.status_code}, {response.text}")

def format_time(iso_time):
    return pd.to_datetime(iso_time)

# Function to calculate time spent between start and end times
def calculate_time_spent(start_time, end_time):
    start = pd.to_datetime(start_time)
    end = pd.to_datetime(end_time)
    time_spent = (end - start).total_seconds() / 60  # Time spent in minutes
    # return format_time(time_spent)
    return f"{int(time_spent // 60)}:{int(time_spent % 60):02d}"  # Return formatted HH:MM
    
def get_first_day_of_month_in_last_paid_invoice_date(project_id, max_retries=3, backoff_factor=2):
    url = f"{BASE_URL}/projects/{project_id}/invoices/"
    retries = 0
    
    while retries < max_retries:
        try:
            response = requests.get(url, headers=headers, auth=HTTPBasicAuth(USERNAME, PASSWORD))
            if response.status_code == 200:
                invoices = response.json().get('invoices', [])
                # paid_invoices = [inv for inv in invoices if inv['status'] == 'paid']
                if invoices:
                    # Find the invoice with the latest date
                    latest_invoice = max(invoices, key=lambda inv: inv['invoiceddate'])
                    latest_invoiced_date = latest_invoice['invoiceddate']

                    # Convert to datetime if it's a string
                    if isinstance(latest_invoiced_date, str):
                        latest_invoiced_date = datetime.strptime(latest_invoiced_date, "%Y-%m-%dT%H:%M:%S")

                    # Set to the first day of the month
                    first_day_of_month = latest_invoiced_date.replace(day=1)

                    # Format as 'YYYY-MM-DDTHH:MM:SS'
                    formatted_date = first_day_of_month.strftime("%Y-%m-%dT%H:%M:%S")

                    print(f"Last paid invoice for project {project_id} found: {latest_invoiced_date} (adjusted to {formatted_date})")
                    return formatted_date

                return None
            elif response.status_code >= 500:
                # Server error, so we should retry
                retries += 1
                wait_time = backoff_factor ** retries
                print(f"Server error {response.status_code} for project {project_id}. Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
            else:
                # Client-side error (e.g., 404), don't retry
                print(f"Client error {response.status_code} for project {project_id}: {response.text}")
                return None
        except requests.exceptions.RequestException as e:
            # Catch network-related errors and retry
            retries += 1
            wait_time = backoff_factor ** retries
            print(f"Network error for project {project_id}: {e}. Retrying in {wait_time} seconds...")
            time.sleep(wait_time)
    
    # If retries are exhausted, log the failure and return None
    print(f"Failed to get invoices for project {project_id} after {max_retries} retries. Skipping project.")
    return None

# Calculate the start and end dates for the previous month
def get_previous_month_dates():
    today = datetime.today()
    first_day_this_month = datetime(today.year, today.month, 1)
    last_day_previous_month = first_day_this_month - timedelta(days=1)
    first_day_previous_month = datetime(last_day_previous_month.year, last_day_previous_month.month, 1)
    return first_day_previous_month, last_day_previous_month

# Function to process both time totals and task details, including all records if a project has records in the previous month
def process_time_per_contact(trackedfrom, trackedto):
    staff_contacts = get_staff_contacts()
    project_data_tasks = {}

    # Get the date range for the previous month
    prev_month_start, prev_month_end = get_previous_month_dates()
    print(f"Including projects with time records from {prev_month_start.strftime('%b %d, %Y')} to {prev_month_end.strftime('%b %d, %Y')}...")

    print("Calculating time per staff member...")

    for contact in staff_contacts:
        contact_id = contact['id']
        contact_name = f"{contact['firstname']} {contact['lastname']}"

        # Get task details per project for the specific contact
        task_details = get_contact_task_details(contact_id, trackedfrom, trackedto)

        # Dictionary to store time records for each project, used to determine if any records fall in the previous month
        project_time_records = {}

        # Process task details per project and remove entries with end_time before last paid invoice
        for record in task_details:
            project_id = record['projectid']
            project_name = record['projecttitle']
            project_number = record['projectnumber']

            # Get the last paid invoice date for the project
            last_paid_invoice_date = get_first_day_of_month_in_last_paid_invoice_date(project_id)
            if last_paid_invoice_date:
                last_paid_invoice_date = pd.to_datetime(last_paid_invoice_date)
            else: last_paid_invoice_date = pd.to_datetime(record['starttime'])

            # # If no last invoice found, set to default
            # if last_paid_invoice_date is None:
            #     last_paid_invoice_date = record['starttime']

            # Check if the task's end_time is after the last paid invoice date
            end_time = pd.to_datetime(record['endtime'])
            if end_time < last_paid_invoice_date:
                print(f"Removing task '{record['taskname']}' for project '{project_name}' because its end time is before the last paid invoice date.")
                continue

            # Store the record in the project's time records
            if project_name not in project_time_records:
                project_time_records[project_name] = []

            project_time_records[project_name].append(record)

        # Check if any time records for this project fall within the previous month
        for project_name, records in project_time_records.items():
            has_prev_month_records = any(
                prev_month_start <= pd.to_datetime(record['endtime']) <= prev_month_end
                for record in records
            )

            # Only include this project if it has records in the previous month
            if has_prev_month_records:
                # Add all records (including older un-invoiced ones) for this project to project_data_tasks
                for record in records:
                    task_name = record['taskname']
                    start_time = record['starttime']
                    task_date = pd.to_datetime(start_time.split('T')[0]).strftime('%b %d, %Y')
                    notes = record.get('notes', '')
                    time_spent = calculate_time_spent(start_time, record['endtime'])
                    formatted_start_time = format_time(start_time)
                    formatted_end_time = format_time(record['endtime'])

                    if project_name not in project_data_tasks:
                        project_data_tasks[project_name] = []

                    project_data_tasks[project_name].append({
                        'Project Name': project_name,
                        'Project Number': record['projectnumber'],
                        'Task Name': task_name,
                        'Task Date': task_date,
                        'Staff': contact_name,
                        'Time Record': notes,
                        'Start': formatted_start_time,
                        'Finish': formatted_end_time,
                        'Time Spent': time_spent
                    })
            else:
                print(f"Skipping project '{project_name}' as it has no time records in the previous month.")

    return project_data_tasks

# Define a function to add color formatting based on column names
def color_headers(worksheet, header_row=2):
    # Define the color mapping for each column name
    color_mapping = {
        "Project Name": "FFC9F8EA",  # Turquoise
        "Project Number": "FFC9F8EA",  # Turquoise
        "Task Name": "FFD7E2FF",  # Blue
        "Task Date": "FFFEEFB8",  # Yellow
        "Staff": "FFFEEFB8",  # Yellow
        "Time Record": "FFFEEFB8",  # Yellow
        "Start": "FFFEEFB8",  # Yellow
        "Finish": "FFFEEFB8",  # Yellow
        "Time Spent": "FFFEEFB8",  # Yellow
    }

    # Iterate over each cell in the header row and apply the color
    for col_idx, cell in enumerate(worksheet[header_row], start=1):
        column_name = cell.value
        color = color_mapping.get(column_name)
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            bold_font = Font(bold=True)
            cell.fill = fill
            cell.font = bold_font
    
    # Create a bold font object
    bold_font = Font(bold=True)

    # Iterate over each cell in the header row and apply the color and bold font
    for col_idx, cell in enumerate(worksheet[header_row], start=1):
        column_name = cell.value
        color = color_mapping.get(column_name)
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.fill = fill
        cell.font = bold_font  # Apply the bold font to the cell

# Define a function to set the font for the entire worksheet
def set_font(worksheet, font_name="Calibri", font_size=10):
    font = Font(name=font_name, size=font_size)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = font

# Function to add the grouped header row
def add_grouped_headers(worksheet):
    # Define the column ranges for each group
    project_columns = ["Project Name"]
    task_columns = ["Task Name"]
    time_record_columns = ["Task Date", "Staff", "Time Record", "Start", "Finish", "Time Spent"]

    # Get the column letters for each group
    project_start = get_column_letter(1)
    project_end = get_column_letter(len(project_columns))
    task_start = get_column_letter(len(project_columns) + 1)
    task_end = get_column_letter(len(project_columns) + len(task_columns))
    time_record_start = get_column_letter(len(project_columns) + len(task_columns) + 1)
    time_record_end = get_column_letter(len(project_columns) + len(task_columns) + len(time_record_columns))

    # Define fills and fonts for each group header
    project_fill = PatternFill(start_color="FFC9F8EA", end_color="FFC9F8EA", fill_type="solid")  # Turquoise
    task_fill = PatternFill(start_color="FFD7E2FF", end_color="FFD7E2FF", fill_type="solid")  # Blue
    time_record_fill = PatternFill(start_color="FFFEEFB8", end_color="FFFEEFB8", fill_type="solid")  # Yellow
    bold_font = Font(bold=True)

    # Merge cells for each group, set their titles, apply alignment, fill, and bold font
    worksheet.merge_cells(f"{project_start}1:{project_end}1")
    project_cell = worksheet[f"{project_start}1"]
    project_cell.value = "PROJECT"
    project_cell.alignment = Alignment(horizontal='center', vertical='center')
    project_cell.fill = project_fill
    project_cell.font = bold_font

    worksheet.merge_cells(f"{task_start}1:{task_end}1")
    task_cell = worksheet[f"{task_start}1"]
    task_cell.value = "TASK"
    task_cell.alignment = Alignment(horizontal='center', vertical='center')
    task_cell.fill = task_fill
    task_cell.font = bold_font

    worksheet.merge_cells(f"{time_record_start}1:{time_record_end}1")
    time_record_cell = worksheet[f"{time_record_start}1"]
    time_record_cell.value = "TIME RECORD"
    time_record_cell.alignment = Alignment(horizontal='center', vertical='center')
    time_record_cell.fill = time_record_fill
    time_record_cell.font = bold_font

# Function to hide specific columns in the worksheet
def hide_columns(worksheet, columns_to_hide):
    for col_name in columns_to_hide:
        # Find the column index based on the column name
        for col_idx, cell in enumerate(worksheet[2], start=1):  # Check in the second row where the headers are
            if cell.value == col_name:
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 0
                break

# Main function to write separate Excel files per project
def main():
    # Process time for all staff contacts
    project_data_tasks = process_time_per_contact(trackedfrom, trackedto)

    # Create output directory if it doesn't exist
    output_dir = 'output/projects/December 2024'
    os.makedirs(output_dir, exist_ok=True)

    # Write each project's data to a separate Excel file
    for project_name, records in project_data_tasks.items():
        df_tasks = pd.DataFrame(project_data_tasks.get(project_name, []))  # Get task details, if available
        
        # Replace invalid characters in project names that can't be used in file names
        safe_project_name = "".join([c if c.isalnum() or c in (' ', '-', '_') else '_' for c in project_name])
        formatted_date = datetime.now().strftime('%b %Y')
        project_number = records[0]['Project Number']
        df_tasks.drop(columns=['Project Number'], inplace=True, errors='ignore')
        excel_file = f'{output_dir}/{project_number} - {safe_project_name} {formatted_date} Timesheet.xlsx'

        # Create pivot table for total time spent by each staff member
        df_tasks['Time Spent in Minutes'] = df_tasks['Time Spent'].apply(lambda x: int(x.split(':')[0]) * 60 + int(x.split(':')[1]))
        pivot_df = df_tasks.pivot_table(index='Staff', values='Time Spent in Minutes', aggfunc='sum').reset_index()

        # Convert the total time spent back to HH:MM format
        pivot_df['Total Time Spent (HH:MM)'] = pivot_df['Time Spent in Minutes'].apply(lambda x: f"{x//60}:{x%60:02d}")

        # Remove the 'Time Spent in Minutes' column as requested in the pivot table
        pivot_df = pivot_df[['Staff', 'Total Time Spent (HH:MM)']]

        # Drop the 'Time Spent in Minutes' from df_tasks to prevent it from appearing in the final Excel output
        df_tasks.drop(columns=['Time Spent in Minutes'], inplace=True)

        # Calculate the sum of total time spent in minutes
        total_time_minutes = pivot_df['Total Time Spent (HH:MM)'].apply(lambda x: int(x.split(':')[0]) * 60 + int(x.split(':')[1])).sum()
        total_time_formatted = f"{total_time_minutes // 60}:{total_time_minutes % 60:02d}"

        # Write DataFrame to Excel
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df_tasks.to_excel(writer, sheet_name='Task Details', index=False, startrow=1)  # Shift down by 1 rows

                # Format "Time Spent" column as duration
            def format_time_spent_as_duration(worksheet, column_letter, start_row, end_row):
                for row in range(start_row, end_row + 1):
                    cell = worksheet[f"{column_letter}{row}"]
                    cell.number_format = '[hh]:mm'  # Format as duration in hours and minutes
            
            # Determine the column letter and row range for "Time Spent" column
            time_spent_col = 'H'  # Adjust if "Time Spent" is in a different column
            start_row = 3  # Row where data starts, adjust as necessary
            end_row = start_row + len(df_tasks) - 1  # Last row of data

            # Write the pivot table to the same sheet, below the task data
            sheet_name = 'Task Details'
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            last_row = len(df_tasks) + 3  # Add 3 for the extra row and header

            # Apply duration formatting to the "Time Spent" column
            format_time_spent_as_duration(worksheet, time_spent_col, start_row, end_row)

            # Add the grouped headers
            add_grouped_headers(worksheet)

            # Apply color formatting to the headers
            color_headers(worksheet)

            # Add a title above the pivot table
            title = "Total Time Per Staff"
            worksheet.cell(row=last_row + 1, column=1, value=title)

            # Write the pivot table below the task details, starting from the title
            for r_idx, row in pivot_df.iterrows():
                for c_idx, value in enumerate(row):
                    worksheet.cell(row=last_row + r_idx + 2, column=c_idx + 1, value=value)

            # Add the total sum below the pivot table
            total_label_row = last_row + len(pivot_df)
            worksheet.cell(row=total_label_row, column=7, value="TOTAL:")
            worksheet.cell(row=total_label_row, column=8, value=total_time_formatted)

            # Adjust the width of each column to fit the content
            for col_idx, col in enumerate(df_tasks.columns, 1):
                max_length = max(df_tasks[col].astype(str).map(len).max(), len(col))  # Find the max length in the column
                worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2  # Add a little padding

            # Apply the font settings to the entire worksheet
            set_font(worksheet, font_name="Calibri", font_size=10)


        print(f"Time data and pivot table for project '{project_name}' saved to {excel_file}")

if __name__ == "__main__":
    main()
